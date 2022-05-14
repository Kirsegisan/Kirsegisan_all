from openpyxl import load_workbook


def inserd_user(*args):
    row = users_page.max_row + 1
    users_page.cell(row=row, column=1).value = args[0]
    users_page.cell(row=row, column=2).value = args[1]
    users_page.cell(row=row, column=3).value = args[2]
    bd.save('data_base.xlsx')


def inserd_sticker(keyword, sticker_id=None, reply_text=None):
    row = stickers_page.max_row + 1
    stickers_page.cell(row=row, column=1).value = keyword
    stickers_page.cell(row=row, column=2).value = sticker_id
    stickers_page.cell(row=row, column=3).value = reply_text
    bd.save('data_base.xlsx')
    stickers[keyword] = sticker_id
    replies[keyword] = reply_text


bd = load_workbook('data_base.xlsx')
stickers_page = bd['stickers']
users_page = bd['users_page']

stickers = {}
replies = {}


def in_database(user: int) -> bool:
    """
    123
    """
    for row in range(2, users_page.max_row + 1):
        if user == users_page.cell(row=row, column=1).value:
            return True
    return False


for row in range(1, stickers_page.max_row + 1):
    keyword = stickers_page.cell(row=row, column=1).value
    sticker_id = stickers_page.cell(row=row, column=2).value
    reply_text = stickers_page.cell(row=row, column=3).value
    stickers[keyword] = sticker_id
    replies[keyword] = reply_text

if __name__ == '__main__':
    print(stickers)
