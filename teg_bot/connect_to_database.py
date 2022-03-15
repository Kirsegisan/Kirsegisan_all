from openpyxl import load_workbook


def inserd_sticker(keyword, sticker_id=None, reply_text=None):
    row = stickers_page.max_row + 1
    stickers_page.cell(row=row, column=1).value = keyword
    stickers_page.cell(row=row, column=2).value = sticker_id
    stickers_page.cell(row=row, column=3).value = reply_text
    bd.save('data_base.xlsx')
    stickers[keyword] = sticker_id
    replies[keyword] = reply_text


bd = load_workbook('data_base.xlsx')
stickers_page = bd['stickers']

stickers = {}
replies = {}

for row in range(1, stickers_page.max_row + 1):
    keyword = stickers_page.cell(row=row, column=1).value
    sticker_id = stickers_page.cell(row=row, column=2).value
    reply_text = stickers_page.cell(row=row, column=3).value
    stickers[keyword] = sticker_id
    replies[keyword] = reply_text

if __name__ == '__main__':
    print(stickers)
